#########################################################
## This program allows one to parse latex to excel.    ##
## Multirow and multicolumn commands will be respected.##
#########################################################

from openpyxl import Workbook
from openpyxl import utils
from openpyxl.styles import Border, Alignment, Side

import click
import re


@click.command()
@click.argument('file', type=click.File('r')) #File is automatically closed
@click.option('-o', 'output_file_name', metavar='<name>', help='output file (less extension)')
def main(file, output_file_name):
    '''
    This script extracts the tables from file and writes the Excel version of them to the file specificed by -o.
    '''

    # Default value for table borders
    double = Side(border_style="thin", color="000000")

    # Specify the location of the upper-left cell of the table
    start_col = 2
    start_row = 2

    # Reads the latex file into memory
    inFile = file.read()
    for idx, table in enumerate(re.findall(r'(\\begin{tabular}.*?\\end{tabular})', inFile, re.DOTALL)):  # Enumerating through all tables in the file
        col_pos_match = re.search(r'\\begin{tabular}{(.*?)}', table)
        col_pos = col_pos_match.group(1)
        total_col = col_pos.count("l") + col_pos.count("c") + col_pos.count("r")
        table = re.sub(r"\\begin{tabular}{.*?}", '', table)
        table = re.sub(r"\\end{tabular}", '', table)
        col = start_col
        row = start_row
        if idx == 0:  # Use the pre-established sheet for table 1
            wb = Workbook()
            ws = wb.active
            ws.title = "Table 1"
        else:  # Creates a new worksheet for all following tables
            ws = wb.create_sheet(title="Table {}".format(str(idx + 1)))
        table = table.split(r"\\")
        for line in table:
            curLine = line

            # Striping away all the special characters in the table
            curLine = re.sub(r"\\cmidrule{.*?}", '', curLine)
            curLine = re.sub(r"\\midrule", '', curLine)
            curLine = re.sub(r"\\toprule", '', curLine)
            curLine = re.sub(r"\\bottomrule", '', curLine)
            curLine = re.sub(r"\\hline", '', curLine)
            curLine = re.sub(r"\\cline{.*?}", '', curLine)
            curLine = curLine.replace("\n", '')
            curLine = curLine.rstrip().split("&")

            # looping through all cells in the current row
            for item in curLine:
                item = item.strip()
                curLoc = utils.cell.get_column_letter(col) + str(row)
                if 'multicolumn' in item:
                    match = re.search(r'\\multicolumn{(.*?)}{(.*?)}{(.*)}', item)
                    text = match.group(3)
                    span2 = 1  # This provides a default in case not nested
                    if 'multirow' in text:
                        match2 = re.search(r'\\multirow{(.*)}{(.*)}{(.*)}', text)
                        if not match2:
                            match2 = re.search(r'\\multirow{(.*)}\[.*\]{(.*)}{(.*)}', text)
                        span2 = int(match2.group(1))
                        # align = match.group(2)
                        text = match2.group(3)
                    span = int(match.group(1))
                    align = match.group(2)
                    try:
                        text = int(text)
                    except:
                        pass
                    if align == "l":
                        align = "left"
                    elif align == "c":
                        align = "center"
                    else:
                        align = "right"
                    ws[curLoc] = text
                    ws[curLoc].alignment = Alignment(horizontal=align, vertical="center")
                    mergeLoc = utils.cell.get_column_letter(col + span - 1) + str(row + span2 - 1)
                    ws.merge_cells("{}:{}".format(curLoc, mergeLoc))
                    col += span
                elif 'multirow' in item:
                    match = re.search(r'\\multirow{(.*)}{(.*)}{(.*)}', item)
                    if not match:
                        match = re.search(r'\\multirow{(.*)}\[.*\]{(.*)}{(.*)}', item)
                    span = int(match.group(1))
                    # align = match.group(2)
                    text = match.group(3)
                    try:
                        text = int(text)
                    except:
                        pass
                    ws[curLoc] = text
                    ws[curLoc].alignment = Alignment(horizontal="center", vertical="center")
                    mergeLoc = utils.cell.get_column_letter(col) + str(row + span - 1)
                    ws.merge_cells("{}:{}".format(curLoc, mergeLoc))
                    col += 1
                else:
                    col += 1
                    if item.replace(" ", "") == "":
                        continue
                    try:
                        item = int(item)
                    except:
                        pass
                    ws[curLoc] = item

            # Applying borders to excel table
            if r"\midrule" in line or r"\toprule" in line or r"\bottomrule" in line or r"\hline" in line:
                loc_1 = utils.cell.get_column_letter(start_col) + str(row  )
                loc_2 = utils.cell.get_column_letter(total_col + start_col - 1) + str(row )
                comb_loc = "{}:{}".format(loc_1, loc_2)
                for cell in ws[comb_loc][0]:
                    cell.border = Border(top=double)
            if r'\cmidrule' in line:
                for match in re.findall(r'\\cmidrule{([0-9]*)-([0-9]*)}', line):
                    loc_1 = utils.cell.get_column_letter(int(match[0]) + start_col - 1) + str(row)
                    loc_2 = utils.cell.get_column_letter(int(match[1]) + start_col - 1) + str(row)
                    comb_loc = "{}:{}".format(loc_1, loc_2)
                    for cell in ws[comb_loc][0]:
                        cell.border = Border(top=double)
            if r'\cline' in line:
                for match in re.findall(r'\\cline{([0-9]*)-([0-9]*)}', line):
                    loc_1 = utils.cell.get_column_letter(int(match[0]) + start_col - 1) + str(row)
                    loc_2 = utils.cell.get_column_letter(int(match[1]) + start_col - 1) + str(row)
                    comb_loc = "{}:{}".format(loc_1, loc_2)
                    for cell in ws[comb_loc][0]:
                        cell.border = Border(top=double)
            col = start_col  # reset column number
            row += 1

    # Saving the workbook
    if output_file_name == None:
        output_file_name=file.name
    wb.save("{}.xlsx".format(output_file_name))
